
#! python3

import numpy as np
import pandas as pd
import os
from openpyxl import load_workbook
from collections import Counter
from datetime import datetime, date, timedelta
from datetime import date
from shutil import copyfile
import time
import re


"""
- rewrite to how too
- what is len(metrics) for? near line 87
- remove all MR


- make sure code works
"""


"""

The purpose of this code is to extract data from an xlsx file, transform it,
then import this new data into a master xlsx file, by searching for missing dates.

Quality is stored in a master excel file, every day auditors extract this data and would
manually review
"""

# removes the annoying pandas warning
pd.options.mode.chained_assignment = None


def date_fixer(df):
    # This function ensures that all date formats are the same
    df['DATE'] = pd.to_datetime(df['DATE'], errors='coerce')
    df['DATE'] = df['DATE'].dt.date

    return df


def diff(li1, li2):
    # Returns the difference of two lists
    return list((set(li1) - set(li2)))


def df_split(push, excel):
    # This will take the new data from two data frames
    holder = pd.DataFrame()
    date_diff = diff(set(push['DATE']), set(excel['DATE']))

    for i in date_diff:
        missing_date = (push[push['DATE'] == i])
        holder = pd.concat([holder, missing_date])

    return holder


################################################################################


# Setting the directory to user's desktop.
desktop = os.path.expanduser("~/Desktop")
os.chdir(desktop)

# location of the master score card
file = (desktop + '/master_file.xlsx')


# Loading master excel file into openpyxl
book = load_workbook(file)
writer = pd.ExcelWriter(file, engine='openpyxl')

writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)


# This is used later to delete data with the udpated edited one. (opening it writer allows to edit the excel)
del_metrics = (writer.sheets['Team_Metrics'])
del_actors = (writer.sheets['Actor_Metrics'])
del_disagree = (writer.sheets['Audit_Data'])

# assigning the sheets into pandas Data Frame (easier to combine with new data)
excel_metrics = pd.read_excel(file, 'Team_Metrics')
excel_actors = pd.read_excel(file, 'Actor_Metrics')
excel_disagreements = pd.read_excel(file, 'Audit_Data')


len_metrics = len(excel_metrics)
len_actors = len(excel_actors)

# these are the new data that needs to be added into the excel spread sheet
sr_disagree = pd.read_csv(desktop + '/audit_data.csv')
metrics = pd.read_csv(desktop + '/team_metrics.csv')
actors = pd.read_csv(desktop + '/actor_metrics.csv')


# Using date_fixer to ensure all dates are in the same format
sr_disagree = date_fixer(sr_disagree)
excel_disagreements = date_fixer(excel_disagreements)
excel_metrics = (date_fixer(excel_metrics))
actors = date_fixer(actors)
excel_actors = date_fixer(excel_actors)
metrics = date_fixer(metrics)


def data_tab(df):

    print("Importing to Data tab")

    #setting DFs by regions from the master excel
    excel_region1 = excel_disagreements[excel_disagreements['region_location']
                                        == 'region 1']
    excel_region2 = excel_disagreements[excel_disagreements['region_location']
                                        == 'region 2']

    #setting DFs by regions from the csv files
    region1 = df[df['region_location'] == 'region 1']
    region2 = df[df['region_location'] == 'region 2']

    # Searches missing date by region, if the Master Excel file has dates missing from the CSV files,
    # it is new data that needs to be added into the Master Excel File.
    # this code will take the missing dates and add them into a new DF
    new_region1 = df_split(region1, excel_region1)
    new_region2 = df_split(region2, excel_region2)

    #Merging the new data's DFs into one DF
    merged_df = pd.concat([new_region1, new_region2])

    # imports the missing dates into the master excel sheet, if there are 0 missing dates
    # nothing will happen
    if merged_df.empty == False:
        merged_df.to_excel(writer, sheet_name='Audit_Data',
                           startrow=writer.sheets['Audit_Data'].max_row, startcol=0, index=False, header=False)
    else:
        pass

    # creating/returning new_df to account for adjustments
    new_df = pd.concat([excel_region1, excel_region2, merged_df], sort=False)
    return new_df

disagree = data_tab(sr_disagree)



print('It has been done Lord Vader')


"""
Since the data from the csv can be manually changed, we need re-aggregate the
data for the metrics tab, while keeping track for histrical data.
"""

#Grouping all the incorrectly labelled jobs so we can count they by date/actor/region
lvl1f_df = disagree[['DATE', 'region_location', 'actor', 'Level_1']]
lvl1f_df = lvl1f_df[lvl1f_df['Level_1'] == 'F']

lvl2f_df = disagree[[
    'DATE', 'region_location', 'actor', 'Level_2', 'content_ID']]
lvl2f_df = lvl2f_df[lvl2f_df['Level_2'] == 'F']

lvl3f_df = disagree[[
    'DATE', 'region_location', 'actor', 'Level_3', 'content_ID']]
lvl3f_df = lvl3f_df[lvl3f_df['Level_3'] == 'F']

lvl4f_df = disagree[[
    'DATE', 'region_location', 'actor', 'Level_4', 'content_ID']]
lvl4f_df = lvl4f_df[lvl4f_df['Level_4'] == 'F']


lvl1F_count_by_actor = lvl1f_df.groupby(['DATE', 'region_location', 'actor', 'Level_1']).size(
).reset_index(name='adjusted_lvl1F').fillna(0)
lvl1F_count_by_region = lvl1f_df.groupby(['DATE', 'region_location', 'Level_1']).size(
).reset_index(name='adjusted_lvl1F').fillna(0)


lvl2F_count_by_actor = lvl2f_df.groupby(['DATE', 'region_location', 'actor', 'Level_2']).size(
).reset_index(name='adjusted_lvl2F').fillna(0)
lvl2F_count_by_region = lvl2f_df.groupby(['DATE', 'region_location', 'Level_2']).size(
).reset_index(name='adjusted_lvl2F').fillna(0)


lvl3F_count_by_actor = lvl3f_df.groupby(['DATE', 'region_location', 'actor', 'Level_3']).size(
).reset_index(name='adjusted_lvl3F').fillna(0)
lvl3F_count_by_region = lvl3f_df.groupby(['DATE', 'region_location', 'Level_3']).size(
).reset_index(name='adjusted_lvl3F').fillna(0)

lvl4F_count_by_actor = lvl4f_df.groupby(['DATE', 'region_location', 'actor', 'Level_4']).size(
).reset_index(name='adjusted_lvl4F').fillna(0)
lvl4F_count_by_region = lvl4f_df.groupby(['DATE', 'region_location', 'Level_4']).size(
).reset_index(name='adjusted_lvl4F').fillna(0)


"""
Importing to the metrics tab
"""
#
print('Importing to OCP_Num_Perc_by_Tier tab')



# Aggregating the disagreements and adding them as new columns in the metrics DF
# this will also account for jons where credit was given back
adjusted = metrics.merge(lvl1F_count_by_region, on=[
                          'DATE', 'region_location'], how='left').fillna(0)

adjusted = adjusted.merge(lvl2F_count_by_region, on=[
                          'DATE', 'region_location'], how='left').fillna(0)
adjusted = adjusted.merge(lvl3F_count_by_region, on=[
                          'DATE', 'region_location'], how='left').fillna(0)

adjusted = adjusted.merge(lvl4F_count_by_region, on=[
                          'DATE', 'region_location'], how='left').fillna(0)



# Creating new columns in the adjusted DF for agreement counts, disagree count and precentage
adjusted['adjusted_lvl1True'] = (
    adjusted['Total'] - (adjusted['adjusted_lvl1F'])).fillna(0)
adjusted['adjusted_lvl2True'] = (
    adjusted['Total'] - (adjusted['adjusted_lvl2F'])).fillna(0)
adjusted['adjusted_lvl3True'] = (
    adjusted['Total'] - (adjusted['adjusted_lvl3F'])).fillna(0)
adjusted['adjusted_lvl4True'] = (
    adjusted['Total'] - (adjusted['adjusted_lvl4F'])).fillna(0)

adjusted['adjusted_lvl1_perc'] = (
    (adjusted['adjusted_lvl1True'] * 1.0) / (adjusted['Total'] * 1.0)).fillna(0)
adjusted['adjusted_lvl2_perc'] = (
    (adjusted['adjusted_lvl2True'] * 1.0) / (adjusted['Total'] * 1.0)).fillna(0)
adjusted['adjusted_lvl3_perc'] = (
    (adjusted['adjusted_lvl3True'] * 1.0) / (adjusted['Total'] * 1.0)).fillna(0)
adjusted['adjusted_lvl4_perc'] = (
    (adjusted['adjusted_lvl4True'] * 1.0) / (adjusted['Total'] * 1.0)).fillna(0)

adjusted['Lvl1_disagree'] = (
    adjusted['Total'] - adjusted['L1_Correct']).fillna(0)
adjusted['Lvl2_disagree'] = (
    adjusted['Total'] - adjusted['L2_Correct']).fillna(0)
adjusted['Lvl3_disagree'] = (
    adjusted['Total'] - adjusted['L3_Correct']).fillna(0)
adjusted['Lvl4_disagree'] = (
    adjusted['Total'] - adjusted['L4_Correct']).fillna(0)

# reorganizing all the columns for adjusted DF and soring them
adjusted = adjusted[['DATE', 'region_location', 'L1_Correct', 'Lvl2_disagree', 'Total', 'L1_Perc', 'L2_Correct', 'Lvl2_disagree', 'Total', 'L2_Perc', 'L3_Correct', 'Lvl3_disagree', 'Total', 'L3_Perc', 'L4_Correct',
                     'Lvl4_disagree', 'Total',	'L4_Perc',	'adjusted_lvl1True', 'adjusted_lvl2True', 'adjusted_lvl3True', 'adjusted_lvl4True',	'adjusted_lvl1_perc', 'adjusted_lvl2_perc', 'adjusted_lvl3_perc', 'adjusted_lvl4_perc']]


adjusted = adjusted.sort_values(by=['DATE', 'region_location'])



# Removing dupliate dates from adjusted DF and Master Excel
date_set = list(set(adjusted['DATE']))
datetime_list = []

for i in date_set:
    datetime_list.append(pd.to_datetime(i))

len_excel_metrics = str(len_metrics + 10)

for rows in del_metrics['A2':('S' + len_excel_metrics)]:
    for i in datetime_list:
        if (rows[0].value) == i:
            del_metrics.delete_rows(rows[0].row)

writer.save()

# adding the adjusted DF to the Master excel file
start_row = []

for rows in del_metrics['A2':('S' + len_excel_metrics)]:

    if rows[0].value is None:
        start_row.append(rows[0].row)
        break

adjusted.to_excel(writer, sheet_name='Metrics', startrow=(
    start_row[0] - 1), startcol=0, index=False, header=False)

writer.save()
#
# """
# Actors Metric's Tab
# """

print("Importing to Accuracy_by_actor_Data tab")

#
adjusted = actors.merge(lvl1F_count_by_actor, on=[
    'DATE', 'region_location', 'actor'], how='left').fillna(0)
adjusted = adjusted.merge(lvl2F_count_by_actor, on=[
                          'DATE', 'region_location', 'actor'], how='left').fillna(0)
adjusted = adjusted.merge(lvl3F_count_by_actor, on=[
    'DATE', 'region_location', 'actor'], how='left').fillna(0)
adjusted = adjusted.merge(lvl4F_count_by_actor, on=[
                          'DATE', 'region_location', 'actor'], how='left').fillna(0)


adjusted['adjusted_lvl1True'] = (
    adjusted['Total'] - (adjusted['adjusted_lvl1F'])).fillna(0)
adjusted['adjusted_lvl2True'] = (
    adjusted['Total'] - (adjusted['adjusted_lvl2F'])).fillna(0)
adjusted['adjusted_lvl3True'] = (
    adjusted['Total'] - (adjusted['adjusted_lvl3F'])).fillna(0)
adjusted['adjusted_lvl4True'] = (
    adjusted['Total'] - (adjusted['adjusted_lvl4F'])).fillna(0)


adjusted['adjusted_lvl1_perc'] = (
    (adjusted['adjusted_lvl1True'] * 1.0) / (adjusted['Total'] * 1.0)).fillna(0)
adjusted['adjusted_lvl2_perc'] = (
    (adjusted['adjusted_lvl2True'] * 1.0) / (adjusted['Total'] * 1.0)).fillna(0)
adjusted['adjusted_lvl3_perc'] = (
    (adjusted['adjusted_lvl3True'] * 1.0) / (adjusted['Total'] * 1.0)).fillna(0)
adjusted['adjusted_lvl4_perc'] = (
    (adjusted['adjusted_lvl4True'] * 1.0) / (adjusted['Total'] * 1.0)).fillna(0)

actor_col = ['DATE', 'actor', 'region_location',	'Total', 'Lvl1_agree',	'Lvl1_disagree', 'Lvl1_Perc', 'Lvl2_agree', 'Lvl2_disagree', 'Lvl2_Perc', 'Lvl3_agree', 'Lvl3_disagree', 'Lvl3_Perc', 'Lvl4_agree', 'Lvl4_disagree', 'Lvl4_Perc',
             'adjusted_lvl1F', 'adjusted_lvl2F', 'adjusted_lvl3F', 'adjusted_lvl4F', 'adjusted_lvl1True', 'adjusted_lvl2True', 'adjusted_lvl3True', 'adjusted_lvl4True', 'adjusted_lvl1_perc', 'adjusted_lvl2_perc', 'adjusted_lvl3_perc', 'adjusted_lvl4_perc']
adjusted = adjusted.reindex(columns=actor_col)

# Aggregating the disagreements and adding them as new columns in the metrics DF
# this will also account for jons where credit was given back
adjusted = adjusted.sort_values(by=['DATE', 'actor'])
adjusted = adjusted.reindex(columns=actor_col)

len_excel_actors = str(len_actors + 10)


# Removing dupliate dates from adjusted DF and Master Excel
date_set = (set(adjusted['DATE']))
datetime_list = []

for i in date_set:
    datetime_list.append(pd.to_datetime(i))

for rows in del_actors['A2':('AC' + len_excel_actors)]:
    for i in datetime_list:
        if (rows[0].value) == i:
            del_actors.delete_rows(rows[0].row)

writer.save()

# adding the adjusted DF to the Master excel file
start_row = []

for rows in del_actors['A2':('R' + len_excel_actors)]:
    if rows[2].value is None:
        start_row.append(rows[0].row)
        break


adjusted.to_excel(writer, sheet_name='Actors', startrow=(
    start_row[0] - 1), startcol=0, index=False, header=False)

writer.save()
writer.close()

print('It has been done Lord Vader')


# # Code written by Tony Nguyen for Facebook, FB PDO team, and Acenture.
